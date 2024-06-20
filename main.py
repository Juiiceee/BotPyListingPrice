import time#librairy pour sleep
import requests#librairy pour faire des requetes
import openpyxl#librairy pour utiliser les tableaux excel
from datetime import date#librairy pas utiliser pour la dte du jour
from bs4 import BeautifulSoup

def change():
    url='https://www.boursorama.com/bourse/devises/convertisseur-devises/dollar-euro'
    html=requests.get(url=url)
    return html.text

def search(query):
    url=f"https://stockx.com/api/browse?_search={query}"
    headers={
        'accept': 'application/json',
        'accept-encoding': 'utf-8',
        'accept-language': 'fr-FR',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1',
        'x-requested-with': 'XMLHttpRequest',
        'app-platform': 'Iron',
        'app-version': '2023.06.04.00',
        'referer': 'https://stockx.com/'
    }
    html = requests.get(url=url,headers=headers)
    return html.json()

wb = openpyxl.load_workbook('requests/shoes.xlsx')
ws=wb['Feuil1']
p=ws['S2':'S40']
y=2
for t in p:
    for i in t:
        value = search(i.value)
        ws['D'+str(y)]=value['Products'][0]['market']['averageDeadstockPrice']
        ws['E'+str(y)]=value['Products'][0]['retailPrice']
        print(value['Products'][0]['title'])
        print("averagePrice : "+str(value['Products'][0]['market']['averageDeadstockPrice']))
        print("RetailPrice : "+str(value['Products'][0]['retailPrice'])+'\n')
        time.sleep(1)
        y+=1
soup=BeautifulSoup(change(),'html.parser')
cl=str(soup.find_all("td",{"class": "c-table__cell c-table__cell--dotted"}))
eur=cl.split()[9]
eur = eur.split(".")
eur = ",".join(eur)
ws['G1']=eur
ws['D41']="=SUMIF(C2:C40,0,D2:D40)"
ws['E41']="=SUM(E2:E40)*G1"
wb.save('requests/hello.xlsx')